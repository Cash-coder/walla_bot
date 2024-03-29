# -*- coding: utf-8 -*-
from multiprocessing import Process, cpu_count
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import traceback
from time import sleep, perf_counter

#load ads to upload
#login in walla using cookies and G user browser
#[upload(ad) for ad in ads]

ADS_FILE      = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\walla_ads_file.xlsx'
ERROR_UPLOADS = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\Not uploaded ads.txt'
# PICS_FOLDER = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\ads_generator\banner_maker\results\results banner girl\\'
#without links
PICS_FOLDER = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\ads_generator\banner_maker\results\old banner no link\\'
#chromdriver
EXECUTABLE_PATH = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\chromedriver.exe'
SITE_URL        = 'https://es.wallapop.com/'
COOKIES_FOLDER  = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\cookies_folder'
BACKUPS_FOLDER  = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\cookies_folder\backups'
PROXY_DATA      = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\login and launcher\proxy_data.py'


#walla_ads_file.xlsx
#Reference = 0
COMPLETE_QUERY_REF = 1
TITLE_COL          = 2
AD_TEXT_COL        = 3
PRICE_COL          = 4
PICS_COL           = 5
CATEGORY           = 6
SUBCATEGORY_1_COL  = 7
SUBCATEGORY_2_COL  = 8
WEIGHT_COL         = 9
PROD_STATE_COL     = 10
BRAND_COL          = 11
MODEL_COL          = 12


# import py file with proxy data from another folder
def importDataFrom(fileName, filePath):
    '''fileName has to include extension like.py 
    filePath has to include filename with extension at the end: x.py'''
    
    from importlib.machinery import SourceFileLoader
    module = SourceFileLoader(fileName, filePath).load_module()

    return module

# import py file with proxy data from another folder
# def import_proxy_data():
#     import sys
#     sys.path.insert(0, PROXY_DATA)
#     import proxy_data
#     return proxy_data


def setFirefoxDriver(proxyN, headless):
    from selenium import webdriver
    from selenium.webdriver.firefox.options import Options
    from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
    from selenium.webdriver.firefox.service import Service

    driverPath      = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\geckodriver.exe'
    profileRootPath = r"C:\Users\HP EliteBook\AppData\Roaming\Mozilla\Firefox\Profiles"

    #proxy settings
    proxy_data = importDataFrom('proxy_data.py', PROXY_DATA)
    HTTPS_PORT = proxy_data.HTTPS_PORT
    IP_ENTRY   = proxy_data.IP_DICT[proxyN]
    IP         = IP_ENTRY.get('ip')
    CITY       = IP_ENTRY.get('city')
    PROFILE    = IP_ENTRY.get('profile')

    print(f"city: {CITY}, IP:{IP}, Profile: {PROFILE}")

    firefox_capabilities = DesiredCapabilities.FIREFOX
    firefox_capabilities['marionette'] = True
    firefox_capabilities['proxy'] = {
    "proxyType": "MANUAL",
    "httpProxy": f'{IP}:{HTTPS_PORT}',
    "sslProxy": f'{IP}:{HTTPS_PORT}'
    }

    #profile settings:
    #to know or create profiles folders
    #search: about:support or about:profiles
    options = Options()
    if headless:
        options.add_argument('-headless')
    options.add_argument("-profile")
    options.add_argument(f"{profileRootPath}{PROFILE}")

    dService = Service(driverPath)
    d = webdriver.Firefox(service=dService, options=options)
    d.maximize_window()

    return d, CITY


def set_driver(proxy_n):
    
    import undetected_chromedriver.v2 as uc
    
    print('setting driver ...')

    proxy_data = importDataFrom('proxy_data.py', PROXY_DATA)
    
    HTTPS_PORT = proxy_data.HTTPS_PORT
    IP_ENTRY   = proxy_data.IP_DICT[proxy_n]
    IP         = IP_ENTRY.get('ip')
    CITY       = IP_ENTRY.get('city')
    PROFILE    = IP_ENTRY.get('profile')

    print(f"city: {CITY}, IP:{IP}, Profile: {PROFILE}")

    options = uc.ChromeOptions()
    options.add_argument(f'--proxy-server=http://{IP}:{HTTPS_PORT}')
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')

    options.add_argument("--headless")
    #to make headless work with cookies load the profile, this'll load the cookies as well
    options.add_argument(r"--user-data-dir=C:\Users\HP EliteBook\AppData\Local\Google\Chrome\User Data")
    options.add_argument(fr'--profile-directory={PROFILE}') 

    d = uc.Chrome(executable_path=EXECUTABLE_PATH, options=options)
    d.maximize_window()
    print('driver set')

    return d, CITY

def load_cookies(d, city):
    import pickle
    from os.path import exists
    from os.path import getsize

    print('loading cookies ...')
    cookies_filepath = f'{COOKIES_FOLDER}\{city}.pkl'
    if not exists(cookies_filepath):
        print(f"This city doesn't exist in the folder: {city} | filepath: {cookies_filepath}")
    if getsize(cookies_filepath) == 0:
        print(f"This cookies file size is 0KB: {cookies_filepath}")
    else:
        # driver has to be in the target URL in order to load the cookies

        d.get(SITE_URL)
        # d.get('https://www.google.com/')

        cookies = pickle.load(open(cookies_filepath, "rb"))
        for cookie in cookies:
            if cookie['domain'] == '.wallapop.com':
                d.add_cookie(cookie)
                # print(f'added this cookie: {cookie}')

        # to apply effect of the cookies reload the page
        print('sleep')
        sleep(7)
        print('refresh')
        # d.refresh()
        print('going to get site')
        d.get('https://www.google.com/')
        print('google got')
        sleep(4)
        d.get(SITE_URL)

        print('cookies loaded')

        return d

def save_cookies(d, city='unspecified'):
    ''' d, city // if city name not specified = cookies_last_login.pkl '''
    import pickle

    print('saving cookies ...')

    if city == 'unspecified':
        city = "cookies_last_login.pkl" 

    # in case new cookies will invalid, backup last used cookies 
    BackupLastUsedCookies(city)
    
    COOKIES_FILE_PATH = COOKIES_FOLDER + '\\' + city + '.pkl'

    with open(COOKIES_FILE_PATH, "wb") as f:
        pickle.dump(d.get_cookies(), f)
    print(f'login cookies saved in city file {city}')

#fiven the city, search that city name in the cookies folder, and copy it to backups adding the date to the filename
def BackupLastUsedCookies(city):
    import os
    import shutil
    import datetime

    flag = 0 #used to flag when there's no file matching name and print the error
    for file in os.listdir(COOKIES_FOLDER):
        file_name = file.split('.')[0]
        if file_name == city:
            flag = 1
            #create a filename with date time in it
            NewName = f'{city} - {datetime.datetime.now()}.pkl'.replace(':', '.') # filenames doesn't accept ":"
            # copy backup, source, destination
            SourceFile  = f'{COOKIES_FOLDER}\{file}'
            Destination = f'{BACKUPS_FOLDER}\{NewName}'
            shutil.copyfile(SourceFile, Destination)
            print(f'created a backup for cookies file {NewName}')
    if flag == 0:    
        print(f'Trying to backup cookies file but no matching city_name was found in backups_folder | city: {city}')

def close_second_tab(d):
    if len(d.window_handles) != 1:
        first_winodw = d.window_handles[0]
        second_winodw = d.window_handles[1]
        
        d.switch_to.window(first_winodw)
        d.close()
        
        # back to where you were
        sleep(1)
        d.switch_to.window(second_winodw)

        print('closed 2º tab')


def load_prods(prods_file):
    # -*- coding: utf-8 -*-
    from openpyxl import load_workbook
   
    wb = load_workbook(filename = prods_file)
    ws = wb.active
    prod_ads = []

    for row in ws.iter_rows(values_only=True, min_row=2):

        if not row[TITLE_COL]:
            continue

        prod_data = {
            'title':   row[TITLE_COL],
            'ad_text': row[AD_TEXT_COL],
            'price':   row[PRICE_COL],
            'pics':    row[PICS_COL],
            'category':    row[CATEGORY],
            'weight':  row[WEIGHT_COL],
            'brand':   row[BRAND_COL],
            'model':   row[MODEL_COL],
            'subcategory_1':  row[SUBCATEGORY_1_COL],
            'subcategory_2':  row[SUBCATEGORY_2_COL],
            'prod_state':     row[PROD_STATE_COL],
                }
        prod_ads.append(prod_data)
    
    return prod_ads

def my_click(d, xpath): 
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait

    wait = WebDriverWait(d, 6)

    # target = wait.until(
        # EC.presence_of_element_located((By.XPATH, xpath)))
        # EC.element_to_be_clickable((By.XPATH, xpath)))
    try:
        #this avoids error: element click intercepted ... Other element would receive the click:
        #select highest level xpath: div inside a header, select header
        target = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))

        target.click()

        #ADVICE: sometimes you can avoid interception error 
        #by sending key RETURN instead of clicking d.xpath.send_keys(Keys.RETURN)

        #Action chains has little accuracy, lots of mistakes clicking in the wrong plage
        # actions = ActionChains(d)
        # actions.move_to_element(target).click().perform()
        # print('clicked')

        sleep(2)

        try:
            d.switch_to.active_element
        except Exception as e:
            #print(e)
            print("can't switch to active element")
    except Exception as e:
        print(f"---Can't click this xpath: {xpath}")
        print(e)

def type_text(d, text, xpath, sleep_time, end_with='unspecified'):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver import ActionChains
    import pyperclip
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait


    wait = WebDriverWait(d, 6)
    element = wait.until(
        # EC.presence_of_element_located((By.XPATH, xpath)))
        EC.element_to_be_clickable((By.XPATH, xpath)))

    #this avoids error: element click intercepted ... Other element would receive the click:
    #select highest level xpath: div inside a header, select header
    wait.until(EC.visibility_of_element_located)
    element.send_keys(text)

    # element = d.find_element(By.XPATH, xpath)
    # element.send_keys(text)

    sleep(sleep_time)

    # element.click()
    # d.execute_script("arguments[0].click;", element)
    # pyperclip.copy(text)
    # act = ActionChains(d)
    # act.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
    
    if end_with == 'tab':
        element.send_keys(Keys.TAB)
    elif end_with == 'enter':
        element.send_keys(Keys.RETURN)
    elif end_with == 'escape':
        element.send_keys(Keys.ESCAPE)

# chromedriver doesn't support emojies, so you use pyperclip to copy text to clipboard? and then Control+V in the selected element to paste
def type_text_with_emoji(d, text, xpath, sleep_time):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver import ActionChains
    import pyperclip

    element = d.find_element(By.XPATH, xpath)
    element.click()
    element.send_keys(text)

    # paste this way if using chromedriver
    # pyperclip.copy(text)
    # act = ActionChains(d)
    # act.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()

    sleep(sleep_time)

def select_category(d, category, subcat1, subcat2):
    
    #click in category drop down
    sleep(2)
    my_click(d, '//tsl-dropdown[@id="category"]')
    sleep(2)

    #in the dropdown menu click element that has text from category parameter
    my_click(d, f'//div[contains(text(), "{category}")]')
    sleep(2)
    #fold back dropdowm menu by clicking something random
    # my_click(d, '//h2[contains(text(), "Información de tu producto")]')
    my_click(d, '//label[@for="price"]')
    
    #click subcategory dropdown menu
    my_click(d, '//tsl-dropdown[@placeholder="Subcategoría"]')

    #click in matching subcategory
    my_click(d, f'//span[contains(text(), "{subcat1}")]')
    sleep(2)

    #fold back the dropdown menu
    my_click(d, '//label[@for="price"]')
    # my_click(d, '//h2[contains(text(), "Información de tu producto")]')

    #sometimes prods have subcat2
    if subcat2:
        #click dropdown menu
        my_click(d, '//tsl-dropdown[@id="objectType2"]')
        sleep(2)
        #click in matching cateorgy
        my_click(d , f'//span[contains(text(), "{subcat2}")]')
        sleep(2)
        #fold back the dropdown menu
        my_click(d, '//label[@for="price"]')


def select_prod_state(d, prod_state):

    #click in prod state dropdown
    my_click(d, '//tsl-dropdown[@id="conditions"]')
    sleep(3)

    # click in dropdown element that matches target prod_state
    my_click(d, f'//li[contains(text(), "{prod_state}")]')
    sleep(2)
    
def insert_pics(d, pic_names, pics_folder):
    
    #locate all 10 pic input paths
    pic_inputs = d.find_elements(By.XPATH, '//input[@type="file"]')

    #wallapop has 10 different pic input paths, one for each pic, so zip each pic with its unique input
    zipped_list = zip(pic_names, pic_inputs)
    
    #insert each pic_local_path into its wallapop_input_path
    for pic_name, input in zipped_list:
        try:

            pic_local_path = pics_folder + pic_name
            # print(f'input: {input} \npic: {pic_local_path}')
            input.send_keys(pic_local_path)
        except Exception as e :
                print(' error in insert_pics(): ', e)
                traceback.print_exc()


def upload_ad(d, ad_data, city):

    # unpack ad_data
    title=    ad_data.get('title')
    ad_text=  ad_data.get('ad_text')
    price=  ad_data.get('price')
    pics=   ad_data.get('pics').split(',')
    weight= ad_data.get('weight')
    brand=  ad_data.get('brand')
    model=  ad_data.get('model')
    category=       ad_data.get('category')
    subcategory_1=  ad_data.get('subcategory_1')
    subcategory_2=  ad_data.get('subcategory_2')
    prod_state= ad_data.get('prod_state')

    sleep(1)
    # click in upload new ad button
    if d.current_url != 'https://es.wallapop.com/app/catalog/upload' :
        d.get('https://es.wallapop.com/app/catalog/upload')
        # my_click(d, '//span[contains(text(), "Subir producto")]')
        sleep(4)

    #click in "Algo que ya no necesito"
    my_click(d, '//span[contains(text(), "Algo que ya no necesito")]')

    # type prod title
    type_text(d, title ,'//input[@id="headline"]', sleep_time=2, end_with='tab')
    
    # price
    type_text(d, price ,'//input[@id="price"]', sleep_time=2, end_with='enter')
    
    #type ad text with emoji
    type_text_with_emoji(d, ad_text ,'//textarea[@id="tellUs"]', sleep_time=2)
    
    #click in cat and sub_cat
    select_category(d, category, subcategory_1, subcategory_2)
    # sometimes the dropdown menu doesn't fold back again, 
    # covering the next click area
    my_click(d, '//label[@for="price"]')

    #click and select prod state
    select_prod_state(d, prod_state)
    # fold back the dropdown menu
    my_click(d, '//label[@for="price"]')

    #insert pics
    insert_pics(d, pics, PICS_FOLDER)

    # type brand | many products don't have brand-model
    if brand:
        type_text(d, brand ,'//input[@placeholder="P. ej: Apple"]', sleep_time=2)
        # specialClickPlaceholderMenu(d, brand)
    # same for model
    if model:
        type_text(d, model, '//input[@placeholder="P. ej: iPhone"]', sleep_time=2)
        # specialClickPlaceholderMenu(d, model)

    #click in prod shipping weight | 2, 5, 10, 20, 30 kg
    # my_click(d, f'//span[contains(text(), "{weight}")]')
    d.find_elements(By.XPATH , '//label[@class="WeightSelector btn ng-star-inserted"]/input')[0].click()

    #LOCATION   
    #upload manually a product with the wanted location, after this Wallapop will remember your selection for future ads

    #click in upload ad
    # my_click(d, '//button[@type="submit"]')
    submitButton = d.find_element(By.XPATH, '//button[@type="submit"]')
    submitButton.send_keys(Keys.ENTER)

    #check if congratulations banner is present to record the uploaded ad to adsDb
    sleep(3)
    try:
        congratBanner = '//button[@class="btn btn-primary btn-primary--bold w-100"]'
        
        #click congratultions banner and button
        my_click(d, congratBanner)
        recordUploadedProd(title, city)
        print(f'successfuly uploaded prod: {title}')
    except Exception as e:
        print('saving screenshot',e)
        d.save_screenshot(f"screenshot: {title}.png")
        recordItemNotUploaded(ad_data, city)


def recordUploadedProd(title, city):
    pass
    'class="info-title subtitle"'

def recordItemNotUploaded(ad, city):
    title = ad.get('title')

    with open(ERROR_UPLOADS, 'a') as f:
        f.write(f'{title}-- {city}\n')

def backupAndClearFileNotUploadedProds():
    import shutil
    import os
    import datetime
    dateHour = datetime.datetime.now().strftime("%d,%B,%H")
    
    SourceFile  = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\NOT uploaded ads.txt'
    Destination = rf'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\not uploaded prods log\NOT uploaded ads -- {dateHour}.txt'

    #if file > 0KB then move to notUploaded log folder 
    fileSize = os.path.getsize(SourceFile)
    if fileSize > 0:
        shutil.move(SourceFile, Destination)

        #create new empty file
        with open('NOT uploaded ads.txt', 'w') as f:
            f.close()

def closePrivacyBanner(d):

    try:
        banner = d.find_element(By.XPATH, '//button[@id="onetrust-accept-btn-handler"]')
        if banner:
            banner.click()
    except Exception as e:
        print(e)
        pass

def run(proxy_n, headless):
        
    ads_to_upload = load_prods(ADS_FILE)
    
    backupAndClearFileNotUploadedProds()

    # d, city = set_driver(proxy_n)
    d, city = setFirefoxDriver(proxy_n, headless)
    # closePrivacyBanner(d)
    sleep(3)

    #click cookies tab
    # my_click(d, '//button[@id="onetrust-accept-btn-handler"]')

    for ad in ads_to_upload:
        
        #if no price then prod not available
        if ad['price'] == None:
            print(f"this ad:{ad['title']} doesn't has price")
            continue
        try:
            print('---uploading: ', ad['title'])

            upload_ad(d, ad, city)            

        except Exception as e:
            recordItemNotUploaded(ad, city)
            print('---------EXCEPTION UPLOADING AN AD!: ', e)
            traceback.print_exc()
            print('going to refresh the page to start a fresh new ad')
            sleep(8)
            #refresh page to clean it and start from 0 a new ad
            d.refresh()
    
    # save_cookies(d, city)
    d.close()

def main():

    headless = True
    a = Process(target=run, args=(0, headless))
    b = Process(target=run, args=(1, headless))
    c = Process(target=run, args=(2, headless))
    
    a.start()
    sleep(1)
    b.start()
    sleep(1)
    c.start()
    
    a.join()
    b.join()
    c.join()

    # #test one in isolated way, without headless
    # headless = False
    # run(2, headless)

    print(f'finish time in {perf_counter()} secs')


if __name__ == "__main__":  
    main()






