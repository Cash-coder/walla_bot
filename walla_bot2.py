# -*- coding: utf-8 -*-
#YOU HAVE TO PUBLISH AT LEAST ONE PRODUCT IN THE TARGET CIty YOU WANT THE REST OF THEM. THEN IT WILL REMEMBER AND MARK THAT CITY ALL THE OTHER UPLOADS. IT'S A WORKAROUND
#wallapop es como Amazon, se vende de todo: https://es.wallapop.com/app/user/oferta-365057118-x6qkp0vwgq6y/reviews
#busca algunos de sus productos con b´suqueda por imágen ?
from selenium.webdriver.common import keys
from pass_module import g_accounts
from selenium.common.exceptions import TimeoutException
import json
import traceback
import undetected_chromedriver.v2 as uc
from selenium.webdriver.common.by import By
import selenium
import time
import random
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import logging


# load prods from ADS_DB
#for city in cities:
    #try login cookies
    #else: login google, then walla with google
    #for product in products
        #upload prod

# COOKIES_FILE =   r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\chatbot\walla_chat\cookies_new.pkl'
COOKIES_FOLDER = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\cookies_folder'
# PRODUCTS_FILE = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\walla_bot\prods.json'
PRODS_FILE =     r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\ADS_DB.xlsx"
PUSBLISHED_ADS_RECORDS_FILE = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\published_ads_records.xlsx"
MUTEX_TIMEOUT = 600 #10'


def login_cookies(d, account_mail, google_or_walla, url):
    import pickle
    #disntiguish between google or walla cookies with google_or_walla
    cookies_file = get_cookies_filepath(account_mail, google_or_walla )
    if cookies_file == 'file not found':
        return 'login error'
    try:
        cookies_filepath = COOKIES_FOLDER + '\\' + cookies_file
        # cookies_filepath = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\chatbot\walla_chat\cookies_new.pkl'
        print(f'using cookies_filepath: {cookies_filepath}')
        d.get(url)
        cookies = pickle.load(open(cookies_filepath, "rb"))
        for cookie in cookies:
            if cookie['domain'] == '.wallapop.com':
            # if google_or_walla in cookie['domain']:
                d.add_cookie(cookie)
                print(f'added this cookie: {cookie}')

        sleep(2)
        d.get(url)
        # logging.info('successful login')
        
        random_sleep(4,8)
        #CHECK LOGIN SUCCESS DEACTIVATED
        #check if login was successful, save cookies
        # if google_or_walla == 'walla':
        #     r = check_presence(d, '//button[@class="BtnLogin"]')
        #     if r == 'not present':
        #         save_cookies(d, COOKIES_FILE=cookies_file)
        #         print('logged in successfully')
        #         return 'success'
        #     elif r == 'present':
        #         return 'login error'
            #if login button not present = loged successfuly
            # try:
            #     login_button_xpath = '//button[@class="BtnLogin"]'
            #     login_button = d.find_element(By.XPATH, login_button_xpath).text
            #     if login_button:
            #         return 'login error'
            # except: #if not found exception = not present = login correct 
            #     save_cookies(d, COOKIES_FILE=cookies_file)
            #     print('logged in successfully')
            #     return 'success'
    except TimeoutException: #sometimes takes too long to load the page, if reaload the url it work
        print('selenium TimeOut exceptiopn, going to reaload the page')
        d.get(url)
        #CHECK LOGIN SUCCESS DEACTIVATED
        #sometimes after loading cookies walla takes too long to load the new page and the old xpaths elements are still present = false bad login , if reaload the apge, those xpaths dissapear = login ok
        # if google_or_walla == 'walla':
        #     r = check_presence(d, '//button[@class="BtnLogin"]')
        #     if r == 'not present':
        #         save_cookies(d, COOKIES_FILE=cookies_file)
        #         print('logged in successfully')
        #         return 'success'
        #     elif r == 'present':
        #         return 'login error'
        
        #CHECK LOGIN SUCCESS DEACTIVATED
        # #if login success, login button shouldn't be present
        # login_button = '//button[@class="BtnLogin"]'
        # r = check_presence(d, login_button)
        # if r == 'not present':
        #     save_cookies(d, COOKIES_FILE=cookies_file)
        #     print('logged in successfully')
        #     return 'success'
        # elif r == 'present':
        #     return 'login error'

    except Exception as e:
        print(f'error in login_cookies(): {e}')
        traceback.print_exc()
        return 'login error'

def check_presence(d, xpath):
    print(f'checking if xpath is prenset: {xpath}')
    try:
        # login_button_xpath = '//button[@class="BtnLogin"]'
        login_button = d.find_element(By.XPATH, xpath)

        print(f'this element it\'s present {login_button.text}')

        if login_button != None:
            print('xpath is present')
            return 'present'
    except:
        print(f'xpath, is NOT prenset: {xpath}')
        return 'not present'

def get_cookies_filepath(account_mail, google_or_walla):
    import os
    global COOKIES_FOLDER
    
    all_cookie_files = os.listdir(COOKIES_FOLDER)
    for file_path in all_cookie_files:
        print(file_path)
        #filepath= google+mail or walla+mail. Use google_or_walla to distinguish
        if account_mail in file_path and google_or_walla in file_path:
            return file_path
    #if no matches
    print(f'not founded any cookies file for this g_account {account_mail}')
    return 'file not found'

def save_cookies(d, COOKIES_FILE='unspecified'):
    ''' d, COOKIES_FILE // if COOKIES_FILE name not specified = cookies_last_login.pkl '''
    import pickle

    if COOKIES_FILE == 'unspecified':
        COOKIES_FILE = "cookies_last_login.pkl" 
    
    COOKIES_FILE_PATH = COOKIES_FOLDER + '\\' + COOKIES_FILE
    with open(COOKIES_FILE_PATH, "wb") as f:
        pickle.dump( d.get_cookies(), f)
    logging.info(f'login cookies saved in file {COOKIES_FILE}')

def cls(): #clears the idle screen
    import os
    os.system('CLS')

# n_proxy = 0  
# def set_driver_old():
#     from selenium import webdriver
#     import proxy_data #py file
#     global n_proxy
    
#     executable_path=r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\chromedriver.exe'

#     IP_LIST = proxy_data.IP_LIST
#     USER    = proxy_data.USER
#     PASS    = proxy_data.PASS
#     SOCKS5_PORT  = proxy_data.SOCKS5_PORT
#     HTTPS_PORT  = proxy_data.HTTPS_PORT

#     # get the next ip of the list
#     if n_proxy == 10: #reset
#         n_proxy = 0
#     NEXT_IP = IP_LIST[n_proxy]
#     n_proxy +=1

#     options =  webdriver.ChromeOptions()
#     options.add_argument('--proxy-server=%s' % NEXT_IP + ":" + '45785')
#     # d = webdriver.Chrome(chrome_options=options,executable_path=executable_path)
#     d = webdriver.Chrome(options=options,executable_path=executable_path)
#     d.maximize_window()

#     return d

def set_driver():
    import undetected_chromedriver.v2 as uc
    import proxy_data
    # global n_proxy

    #HEADLESS
    # print('setting driver...')
    # options = uc.ChromeOptions()
    # # options.add_argument('--no-sandbox')
    # # options.add_argument('--headless')
    # # options.add_argument('--disable-gpu')
    # # options.add_argument("--window-size=1920, 1200")
    # # options.add_argument('--disable-dev-shm-usage')
    #PROXY
    # IP_LIST = proxy_data.IP_LIST
    # USER    = proxy_data.USER
    # PASS    = proxy_data.PASS
    # SOCKS5_PORT  = proxy_data.SOCKS5_PORT
    # HTTPS_PORT  = proxy_data.HTTPS_PORT
    # # get the next ip of the list
    # if n_proxy == 10: #reset
    #     n_proxy = 0
    # NEXT_IP = IP_LIST[n_proxy]
    # n_proxy +=1
    # print(f'{NEXT_IP}:{HTTPS_PORT}')
    
    executable_path=r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\chromedriver.exe'

    options = uc.ChromeOptions()
    # options.add_argument(f'--proxy-server=socks5://{NEXT_IP}:{SOCKS5_PORT}')
    # options.add_argument(f'--proxy-server=https://{NEXT_IP}:{HTTPS_PORT}')
    # options.add_argument(f'--proxy-server=https://185.121.15.68:45785')
    # options.add_argument(f'--proxy-server=https://{NEXT_IP}:{HTTPS_PORT}')
    # options.add_argument('ignore-certificate-errors')

    d = uc.Chrome(executable_path=executable_path,options=options)
    # d = uc.Chrome(options=options)
    d.set_page_load_timeout(15) #sometimes takes too long to load a page, used with timeout excpetion to reload the page
    d.maximize_window()

    #close 1º window
    if len(d.window_handles) != 1:
        first_winodw = d.window_handles[0]
        second_winodw = d.window_handles[1]
        
        d.switch_to.window(first_winodw)
        d.close()
        sleep(2)
        d.switch_to.window(second_winodw)
        
    print(f'driver set. Session id: {d.session_id}')
    return d


def my_click(d, xpath): 
    # print('going to click:', xpath)
    wait = WebDriverWait(d, 5)
    target = wait.until(
    EC.presence_of_element_located((By.XPATH, xpath)))
    try:
        
        #target.click()
        #Action chains has little accuracy, lots of mistakes clicking in the wrong plage
        actions = ActionChains(d)
        actions.move_to_element(target).click().perform()
        # print('clicked')

        try:
            d.switch_to.active_element
        except Exception as e:
            #print(e)
            print("can't switch to active element")
        
        # sleep(3)
    except Exception as e:
        print(e)

def type_text(d, text, xpath, end='unspecified'):
    ''' d, xpath, text, end: tab, enter'''
    element = d.find_element(By.XPATH, xpath)
    element.send_keys(text)
    
    if end == 'tab':
        element.send_keys(Keys.TAB)
    elif end == 'enter':
        element.send_keys(Keys.RETURN)

def select_prod_state(d, prod_state):
    ''' d , prod_state'''
    print(f'in select_prod_state() prod_state: {prod_state}')
    
    #buen estado if not_new, new if new
    if prod_state == 'not_new':
        target_xpath = '//li[contains(text(), "En buen estado")]'
    elif prod_state == 'new':
        target_xpath = '//li[contains(text(), "Nuevo")]'

    try:
        my_click(d, '//div[contains(text(), "Escoge un estado")]')
        my_click(d, target_xpath)
    except Exception as e:
        print(f'error in select_prod_state() : {e}')
        traceback.print_exc()
    # OLD, in desuse
    # try:#open dropdown list
    #     my_click(d, '//div[contains(text(), "Escoge un estado")]')
    #     #select based on prod_state
    #     if prod_state == 'new':
    #         my_click(d, '//li[contains(text(), "Nuevo")]')
    #     elif prod_state == 'like new':
    #         my_click(d, '//li[contains(text(), "Como nuevo")]')
    #     elif prod_state == 'good state':
    #         my_click(d, '//li[contains(text(), "En buen estado")]')
    #     elif prod_state == 'acceptable state':
    #         my_click(d, '//li[contains(text(), "En condiciones aceptables")]')
    #     elif prod_state == 'd state':
    #         my_click(d, '//li[contains(text(), "Lo ha dado todo")]')

def insert_pics(d, pics):
    print(f' in insert_pics()  pics: {pics}')
    # r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Imágenes\4Oscuro.jpg'
    # ["C:\\Users\\HP EliteBook\\OneDrive\\A_Miscalaneus\\Imágenes\\4Oscuro.jpg", "C:\\Users\\HP EliteBook\\OneDrive\\A_Miscalaneus\\Imágenes\\2oscuro.jpg","C:\\Users\\HP EliteBook\\OneDrive\\A_Miscalaneus\\Imágenes\\3oscuro.jpg"]
    #pics = [r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Imágenes\4Oscuro.jpg', r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Imágenes\2oscuro.jpg', r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\iphone-13_2.jpg', r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\iphone13_1.jpg']
    # pic = 'https://mega.co.nz/#!SrBgjByT!xXMd75g13YVk_iAgdcOhAJ5FDZRU4NWDqlyYGGTska8'
    
    pics_input_xpath = '//input[@type="file"]'
    pic_inputs = d.find_elements(By.XPATH, pics_input_xpath)
    
    #wallapop has 10 different pic input paths, one for each pic, so zip each pic with its unique input
    zipped_list = zip(pics, pic_inputs)
    
    for entry in zipped_list:
        try:
            pic = entry[0]
            input = entry[1]

            input.send_keys(pic)
            # print(f'pic: {pic}')
            # print(f'input: {input}')
            print(f'inserted pic {pic}')
        except Exception as e :
            print(' error in insert_pics(): ', e)
            traceback.print_exc()

def upload_ad(  d,
                prod_title,
                prod_state,
                prod_description,
                price,
                prod_brand,
                prod_model,
                walla_location,
                pics,
                category,
                subcategory = 'not needed'
                ):
    
    d.get('https://es.wallapop.com/app/catalog/upload')

    # print(f'in upload_ad() \n prod_title {prod_title} \n prod_state {prod_state} \n prod_description {prod_description} \n price {price} \n prod_brand {prod_brand} \n prod_model {prod_model} \n location {location} \n pics {pics} \n category {category}')

    try:
        # if driver not in upload ad page, click in upload new ad
        if d.current_url != 'https://es.wallapop.com/app/catalog/upload' :
            my_click(d, '//span[contains(text(), "Subir producto")]')
        
        sleep_time = 1

        #click in "algo que ya no necesito"
        my_click(d, '//span[contains(text(), "Algo que ya no necesito")]')
        sleep(sleep_time)
        random_sleep(1,3)
        #prod title
        type_text(d, prod_title ,'//input[@placeholder="En pocas palabras..."]')
        sleep(sleep_time)
        random_sleep(1,3)
        #price
        type_text(d, price ,'//input[@placeholder="(No te pases)"]')
        sleep(sleep_time)
        random_sleep(1,3)
        # category
        select_category(d, category, subcategory)        
        sleep(sleep_time)
        random_sleep(1,3)
        #prod state (sin abrir, nuevo, como nuevo, en buen estado, en condiciones aceptables, lo ha dado todo)
        select_prod_state(d, prod_state)
        sleep(sleep_time)
        random_sleep(1,3)
        #fill prod_description
        # type_text(d, prod_description, '//textarea')
        type_slow(d, prod_description, '//textarea')
        random_sleep(1,3)
        #fill brand (iphone)
        prod_brand = create_brand(prod_title)
        brand_input_xpath = '//input[@placeholder="P. ej: Apple"]'
        type_slow(d, prod_brand, brand_input_xpath, end='TAB')
        random_sleep(1,3)
        #type TAB to drop up brands list
        brand_input = d.find_element(By.XPATH, brand_input_xpath)
        brand_input.send_keys(Keys.TAB)
        random_sleep(1,3)
        #fill model (iphone 12)
        ###########################################
        type_slow(d, prod_model, '//input[@placeholder="P. ej: iPhone"]', end='TAB')
        random_sleep(1,3)
        #upload pics, list of absolute paths
        insert_pics(d, pics)     
        random_sleep(1,3)
        #click in shipping 2Kg
        my_click(d, '//span[contains(text(), "2kg")]')
        random_sleep(1,3)
        # type location
        sleep(5)
        location_input = d.find_element_by_xpath('//input[@placeholder="Marca la localización"]')
        location_input.send_keys(Keys.CONTROL,'a', Keys.DELETE)
        location_input.send_keys(walla_location)
        location_input.send_keys(Keys.RETURN)
        # iframe = d.find_element_by_xpath('//iframe[@id="tdz_ifrm"]')
        # d.switch_to.frame(iframe)
        
        # type_slow(d, walla_location, '//input[@placeholder="Marca la localización"]')
        sleep(8)
        # my_click(d, '//button[contains(text(), "Aplicar")]')
        # apply_location_button = d.find_element_by_xpath('//button[contains(text(), "Aplicar")]')
        apply_location_button = d.find_element(By.XPATH, '//button[contains(text(), "Aplicar")]')
        apply_location_button.click()
        # apply_location_button.click()
        # actions = ActionChains(d)
        # actions.move_to_element(apply_location_button).click().perform()
        print('done')

        random_sleep(20,45)

        # d.switch_to.parent_frame()

        #click in submit
        my_click(d, '//button[@class="btn btn-block btn-primary"]')
        random_sleep(1,3)
        
        #click in close "Genial tu producto ya está subido, queres destacarlo ?"
        my_click(d , '//div[@class="modal-close light"]')
        random_sleep(1,3)
        
        # '//tsl-svg-icon[@src="/assets/icons/cross.svg"]'
        random_sleep(5,10)
        try:
            d.switch_to.active_element #switch to uploaded panel
            success_text = d.find_element(By.XPATH, '//h1[contains(text(), "¡Genial! Tu producto ya está en Wallapop")]').text

            # success_text_xpath = '//h1[contains(text(), "¡Genial! Tu producto ya está en Wallapop")]'
            
            # wait = WebDriverWait(d, 10)
            # success_text = wait.until(
            # EC.presence_of_element_located((By.XPATH, success_text_xpath)))

            if success_text != None:
                print(f'successful upload {prod_title}')
                return 'success'
        except Exception as e:
            print(e)
            traceback.print_exc()
            print(f'not uploaded this ad {prod_title}')
            return 'uploading error'

    except Exception as e:
        print(f'---error in upload_ad(), error: {e}')
        traceback.print_exc()
        return 'uploading error'
    
    # if total_published_ads == prod_len-1: # -1 because the 1º is credentials
    #     print('published all ads, driver closed')           
    #     d.close()

def select_category(d, category, subcategory='not needed'):
    print(f'category inside select_category(): {category}')
    if category == 'smartphones':
        my_click(d, '//tsl-dropdown[@placeholder="Categoría"]')
        my_click(d, '//div[contains(text(), " Móviles y Telefonía ")]') #notice it has spaces at both sides
        my_click(d, '//tsl-dropdown[@placeholder="Subcategoría"]')
        my_click(d, '//span[contains(text(), "Teléfonos móviles")]')

def create_brand(prod_tile):
    try:
        prod_tile = prod_tile.lower()
        if 'iphone' in prod_tile:
            prod_brand = 'Apple'
        elif 'samsung' in prod_tile:
            prod_brand = 'Samsung'
        elif 'xiaomi' in prod_tile:
            prod_brand = 'Xiaomi'
        elif 'google pixel' in prod_tile:
            prod_brand = 'google pixel'
        elif 'oppo' in prod_tile:
            prod_brand = 'Oppo'
        elif 'realme' in prod_tile:
            prod_brand = 'Realme'
        elif 'oneplus' in prod_tile:
            prod_brand = 'Oneplus'

        return prod_brand

    except Exception as e:
        print('in create_brand(): ',e)
        traceback.print_exc()

def my_mutex(file_path):
    from openpyxl import load_workbook
    global MUTEX_TIMEOUT
    #this way the variable sets to MUTEX_TIMEOUT each time the function is called
    mutex_timeout = MUTEX_TIMEOUT

    while mutex_timeout > 0:
        wb = load_workbook(filename = file_path)
        ws = wb.active
        
        mutex_flag = ws['A1'].value
        print(mutex_flag)
        if mutex_flag == 'file free to access':
            wb.close()
            return 'file free to access'
        elif mutex_flag == 'occupied':
            sleep(1)
            print('file occupied')
            mutex_timeout += -1
    if mutex_timeout == 0:
        wb.close()
        return 'file occupied mutex timeout'

def set_mutex(state, file_path):
    from openpyxl import load_workbook
    wb = load_workbook(filename = file_path)
    ws = wb.active

    ws['A1'] = state
    print(f'set_mutex(): file <{file_path}> set to {state} ')
    wb.save(file_path)
    wb.close()

def load_prods(prods_file):
    #creates a list of dicts
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    wb = load_workbook(filename = prods_file)
    ws = wb.active

    all_prods = []
    for row in ws.iter_rows(values_only=True, min_row=4):
        id    =       row[0]
        target_title= row[1]
        prod_state  = row[2]
        walla_text  = row[3]
        #reserved for mila text [4]
        add_price   = row[5]
        pic_path_1  = row[6]
        pic_path_2  = row[7]
        pic_path_3  = row[8]
        pic_path_4  = row[9]
        target_cateory=row[10]
        target_model = row[11]

        prod = [id, target_title, prod_state, walla_text, add_price, pic_path_1, pic_path_2, pic_path_3, pic_path_4, target_cateory, target_model]
        all_prods.append(prod)

    return all_prods

def type_slow(d,text,xpath, end='key'):
    ''' d, text, xpath // given driver, text and xpath input it types the text in the xpath input'''
    import time
    import random

    try: #it gets a weird excpetion, probably some python bug
        elem = d.find_element_by_xpath(xpath)
        for character in text:
            # print(f'going to write this : {character} ')
            elem.send_keys(character)
            time.sleep(random.uniform(0.05,0.08))
        
        if end == 'TAB': #return, tab
            elem.send_keys(Keys.TAB)

    except selenium.common.exceptions.ElementNotInteractableException:
        print('not interactable exception, passing')
        pass

def random_sleep(a, b):
    from random import randint
    sleep(randint(a, b))

def press_enter(d):
    actions = ActionChains(d)
    actions.send_keys(Keys.ENTER)
    actions.perform()

def google_human_login(d, account_email, account_pasword):

    #try to load google cookies
    #if any, use login form
    #save cookies google+account_mail

    print('google login, trying to load cookies')
    google_cookies = get_cookies_filepath(account_email, 'google')
    r = login_cookies(d, account_email, 'google', 'https://www.google.com/')
    print('cookies loaded successfully, going to save cookies')
    if r == 'success':
        cookies_filename = 'google-'+account_email+'.pkl'
        save_cookies(d, cookies_filename)
        print('cookies saved successfully')
        return 'success' #return is used to stop the funtion here, 'success' is not used

    d.get('https://www.google.com/')
    
    #switch to search tab
    if len(d.window_handles) > 1: #sometimes it opens 2 tabs, one with google chrome start account
        d.switch_to.window(d.window_handles[1])
    
    #accept cookies
    try:
        my_click(d,'//div[contains(text(),"Acepto")]')
    except:
        print('accept cookies banner')
        pass

    login_link = 'https://accounts.google.com/signin/chrome/sync/identifier?ssp=1&continue=https%3A%2F%2Fwww.google.com%2F&flowName=GlifDesktopChromeSync'    
    d.get(login_link)
    # type email
    email_xpath = "//input[@type='email']"
    type_slow(d, account_email, email_xpath)
    sleep(2)
    press_enter(d)
    sleep(100)
    #type password
    password_xpath = '//input[@type="password"]'    
    type_slow(d, account_pasword, password_xpath)
    press_enter(d)
    sleep(8)
    

    cookies_filename = 'google_'+account_email+'.pkl'
    save_cookies(d, cookies_filename)
    print(f'saved cookies {cookies_filename}')

def walla_google_login(d, account_mail):

    d.get('https://es.wallapop.com/')

    #cookies_agreement
    try:
        my_click(d,'//button[@id="didomi-notice-agree-button"]')
    except:
        logging.info('No coookies agreement button')
        pass
    #login button
    my_click(d,'//button[@class="BtnLogin"]')
    sleep(4)
    #entrar con google
    #my_click(d,"//a[contains(text(), 'Inicia sesión')]")
    my_click(d,'//button[contains(text(),"Entrar con Google")]')
    print('sleeping to wait walla_google login')
    sleep(15)

    try: #sometimes it asks to pick an user, other times uses your current loged user
        #switch to accounts options dropdown list and click option with account_email
        d.switch_to.window(d.window_handles[1])
        # if 'https://accounts.google.com/o/oauth2/' in d.current_url:
        xpath = f'//div[@data-email="{account_mail}"]'
        my_click(d, xpath)
    except Exception as e:
        print('error trying to click account in walla_google_login, probably because walla didn\'t ask to pick a user',e)
        pass

    #detect something that asserts you're logged in
    random_sleep(2,5)
    try:
        #assert successful login
        r = check_presence(d, '//button[@class="BtnLogin"]')
        #if login button not present = login ok
        if r == 'not present':
            print('login success')
            
            #save success cookies 
            cookies_file_name = 'walla'+account_mail + '.pkl'
            save_cookies(d,  cookies_file_name)

            return 'success'
        
    except Exception as e:
        print(f'in walla_google_login(): {e}')
        traceback.print_exc()

def record_published_ad(id, account_city):
    # to make the code faster, instead of parse all ads_db, just append to a file
    #at then end other program reads the file row by row and updates ads_db file
    from openpyxl import load_workbook
    wb = load_workbook(PUSBLISHED_ADS_RECORDS_FILE)
    ws = wb.active

    ws.append([id, account_city])
    wb.save(PUSBLISHED_ADS_RECORDS_FILE)
    wb.close()

def create_location(account_city):
    account_city = account_city.lower()
    if account_city == 'madrid':
        location = 'Madrid, Madrid'

    if account_city == 'barcelona':
        location = 'barcelona, barcelona'
    if account_city == 'málaga':
        location = 'málaga, málaga'
    if account_city == 'valencia':
        location = 'valencia, valencia'
    if account_city == 'sevilla':
        location = 'sevilla, sevilla'
    if account_city == 'zaragoza':
        location = 'zaragoza, zaragoza'
    if account_city == 'murcia':
        location = 'murcia, murcia'
    if account_city == 'bilbao':
        location = 'bilbao, bilbao'
    if account_city == 'alicante':
        location = 'alicante, alicante'
    if account_city == 'córdoba':
        location = 'córdoba, córdoba'
    if account_city == 'valladolid':
        location = 'valladolid, valladolid'
    if account_city == 'vigo':
        location = 'vigo, vigo'
    if account_city == 'gijón':
        location = 'gijón, gijón'
    if account_city == 'hospitalet':
        location = 'hospitalet, hospitalet'
    if account_city == 'la coruña':
        location = 'La coruña, la coruña'
    if account_city == 'elche':
        location = 'elche, elche'
    if account_city == 'granada':
        location = 'granada, granada'
    if account_city == 'tarrasa':
        location = 'tarrasa, tarrasa'
    if account_city == 'badalona':
        location = 'badalona, badalona'
    if account_city == 'oviedo':
        location = 'oviedo, oviedo'
    if account_city == 'sabadell':
        location = 'sabadell, sabadell'
    if account_city == 'jerez de frontera':
        location = 'jerez de frontera, jerez de frontera'
    if account_city == 'móstoles':
        location = 'móstoles, móstoles'
    if account_city == 'pamplona':
        location = 'pamplona, pamplona'
    if account_city == 'almería':
        location = 'almería, almería'

    return location


def run():
    
    global total_published_ads

    #sould start a new driver in each account?= more stable, if 1 account fails the error don't extend to other accounts
        
    for account in g_accounts[2:]: #imported from pass.py
        account_city = account[0]
        account_mail = account[1]
        google_pass = account[2]
        walla_pass  = account[3]

        print(f'account_city: <{account_city}> account_mail: <{account_mail}> google_pass: <{google_pass}> walla_pass : <{walla_pass}>')

        d = set_driver()
        try:

            # login
            r = login_cookies(d, account_mail, 'walla', 'https://es.wallapop.com/')
            if r == 'login error': # if can't login with cookies, login google adn then in wallapop having the google
                google_human_login(d, account_mail, google_pass)
                r = walla_google_login(d, account_mail) #mail used to name cookies 
                if r != 'success':
                    print('couldn\'t login, or the user_zone xpath is not correct')
                    continue # to next account
            else:
                print('login success')

            # iterate products
            total_published_ads = 0

            products = load_prods(PRODS_FILE)
            prods_len = len(products)
            print(f'total n loaded of products: {len(products)}')

            n_prods = 0 #used to random sleep every x products
            for prod in products:
                try:
                    if n_prods == 3: #upload chunks of 3 and rest
                        random_sleep(20,40)
                        n_prods = 0

                    n = 0
                    for row in prod:
                        print(f' {n} --- {row}')
                        n += 1

                    id    =       prod[0]
                    prod_title  = prod[1]
                    prod_state  = prod[2]
                    prod_description  = prod[3]
                    # 4 if mila_text
                    price       = prod[4]
                    pic_path_1  = prod[5]
                    pic_path_2  = prod[6]
                    pic_path_3  = prod[7]
                    pic_path_4  = prod[8]
                    target_category=prod[9]
                    prod_model = prod[10]
                    prod_model = str(prod_model)
                    # print(f'this is model {prod_model}')

                    if prod_title == None:
                        continue

                    pics = [pic_path_1, pic_path_2, pic_path_3, pic_path_4]
                    prod_brand = create_brand(prod_title)
                    walla_location   = create_location(account_city)
                    
                    r = upload_ad(d,prod_title = prod_title,
                                    prod_state = prod_state,
                                    prod_description = prod_description,
                                    price = price,
                                    prod_brand = prod_brand,
                                    walla_location = walla_location,
                                    pics = pics,
                                    prod_model = prod_model,
                                    category = target_category
                                    )
                    if r == 'uploading error':
                        print('uploading error')
                        continue
                    elif r == 'success':
                        total_published_ads += 1
                        record_published_ad(id, account_city)
                        print(f'total successful published ads: {total_published_ads} of {prods_len} total // new ad published: {prod_title}')
                #upload prod exception
                except Exception as e:
                    print(f'error in loop uploading prods prod_title: {prod_title}, error message: \n {e}')
                    continue
        finally:   
            print(f'published all ads: {total_published_ads}, driver closed')           
            d.close()
            random_sleep(60, 200)

    


if __name__ == "__main__":
    run()